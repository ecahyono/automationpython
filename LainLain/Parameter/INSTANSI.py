# import imp
from json import load
from lib2to3.pgen2 import driver
import string
from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains


from openpyxl import load_workbook
import time

#target halaman excel ada dimana , wb = variablenya
# wb = load_workbook(filename="C:\chromedriver\Data.xlsx")
wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\lainlain.xlsx")

# jadi ini bisa read sheet yang dibawah itu yang di excel
sheetrange = wb['Instansi']

# ini web driver disimpen dimana, kalo disimpen di path kosongin aja
driver = webdriver.Chrome(r'C:\Users\user\Documents\TRCH\chromedriver.exe')

# link nya ini dimana
driver.get("http://kumbang.torche.id:32400/")
# seting windows nya jadi max   
driver.maximize_window()
# script gakan di eksekusi kalo web ga muncul. kalo lebih dari 10 detik ga muncul error
driver.implicitly_wait(10)
# ini letak xpath icon login
driver.find_element(By.XPATH, "//div/span").click()
# ini masuk ke form input username
driver.find_element(By.ID, "username").click()
# masukin input username
driver.find_element(By.ID, "username").send_keys("rono")
# masukin input password
driver.find_element(By.ID, "password").send_keys("rene")
# click button login
driver.find_element(By.ID, "kc-login").click()
time.sleep(5)

element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[7]/div")
actions = ActionChains(driver)
actions.move_to_element(element).perform()
time.sleep(3)

# WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "XPATH LAIN LAIN PARAMETER "))) ## OPSIONAL

element2 = driver.find_element(By.XPATH, "//div[9]/div/ul/li/div")
actions2 = ActionChains(driver)
actions2.move_to_element(element2).perform()
driver.find_element(By.LINK_TEXT, "Instansi").click()

i =2 

while i <= len(sheetrange['A']):
    # deklarasi bahwa NIP itu ada di A 
    Instansi = sheetrange['A'+str(i)].value
    # deklarasi bahwa NAMA itu ada di B 
    Instansi_Diatasnya = sheetrange['B'+str(i)].value
    # deklarasi bahwa NAMA itu ada di B 
    Keterangan = sheetrange['C'+str(i)].value
    
    time.sleep(1)

    driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/div[1]/button").click()

    try:
        WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/div[2]/div/div[2]/div/div/div/div/h1')))
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[1]/div/div[1]/input").send_keys(Instansi)
        # driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div[1]/div[2]/div/div/form/div[2]/div/div/div/div").send_keys(instansi_Diatasnya)
        time.sleep(2)
        driver.find_element(By.XPATH, "(//input[@type=\'text\'])[2]").send_keys(Instansi_Diatasnya)
        time.sleep(6)
        driver.find_element(By.XPATH, "//li[contains(.,'POLDA JABAR')]").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//form/div[3]/div/div/span").click()
                                    
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[8]/div/div/textarea").send_keys(Keterangan)
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/div[2]/div/div[2]/div/div/form/div[9]/div/div/button[2]").click()
        
        
    except TimeoutException:
        # print("d")
        pass
    time.sleep(7)   
    i = i + 1
print ("Success Created")