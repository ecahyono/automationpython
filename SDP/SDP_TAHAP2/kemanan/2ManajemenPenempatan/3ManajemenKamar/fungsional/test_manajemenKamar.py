from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.firefox.options import Options
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))


from Settings.setup import initDriver, loadDataPath, quit, sleep
from Settings.loginkeamanan import loginOperatorSumedang, Op_Keamanan_p2u, SpvRutanBdg, op_keamanan_mp
from Settings.Page.keamanan import manajemenpenghunibaru

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('LogManajemenKamar.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

wb = load_workbook(environ.get("KeamananUAT"))
sheetrange = wb['ManajemenKamarIndex']
i = 2

    
Nama                                    = sheetrange['A'+str(i)].value

Kejahatan                               = sheetrange['B'+str(i)].value
blokform                                = sheetrange['C'+str(i)].value
Lantai                                  = sheetrange['D'+str(i)].value
Kamar                                   = sheetrange['E'+str(i)].value
TanggalPenempatan                       = sheetrange['F'+str(i)].value
Keterangan                              = sheetrange['G'+str(i)].value


@mark.fixture_test()
def test_2_login():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')
    op_keamanan_mp(driver)
    Log.info('Login Operator Manajemen Penempatan')


@mark.fixture_test()
def test_3_AksesMenu():
    sleep(driver)
    manajemenpenghunibaru(driver)
    print('.')
    Log.info('Akses halaman Manajemen Penghuni Baru')
    attach(data=driver.get_screenshot_as_png())
    Log.info('Akses Menu Manajemen Kamar')

@mark.fixture_test()
def test_4_FilterColumn():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    driver.find_element(By.ID, "filterColumn").click()
    driver.find_element(By.ID, "nama_lengkap").click()
    driver.find_element(By.ID, "kataKunci").click()
    driver.find_element(By.ID, "kataKunci").send_keys(Nama)
    Log.info("Search Nama lengkap")

@mark.fixture_test()
def test_5_ClickButtonSearch():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    driver.find_element(By.CSS_SELECTOR, "#searchButton svg").click()
    Log.info('Click Button Search')

@mark.fixture_test()
def test_6_ClickButtonMutasi():
    sleep(driver)
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    driver.find_element(By.ID, "mutasi-0").click()
    Log.info('Click Button mutasi')

@mark.fixture_test()
def test_7_PilihBlok():
    sleep(driver)
    driver.find_element(By.ID, "blokForm").send_keys(blokform)
    driver.find_element(By.ID, "blokOption-0").click()
    Log.info('Pilih Blok')

@mark.fixture_test()
def test_8_PilihLantai():
    sleep(driver)
    driver.find_element(By.ID, "lantaiForm").send_keys(Lantai)
    driver.find_element(By.ID, "lantaiOption-0").click()
    Log.info('Pilih Lantai')

@mark.fixture_test()
def test_9_PilhKamar():
    sleep(driver)
    driver.find_element(By.ID, "kamarForm").send_keys(Kamar)
    driver.find_element(By.CSS_SELECTOR, "#kamarOption-0 > span").click()
    Log.info("Pilih Kamar")

@mark.fixture_test()
def test_10_InputTanggalMutasi():
    sleep(driver)
    driver.find_element(By.ID, "tanggalMutasiForm").click()
    driver.find_element(By.ID, "tanggalMutasiForm").send_keys(TanggalPenempatan)
    Log.info('input tanggal penempatan')

@mark.fixture_test()
def test_11_InputKeterangan():
    sleep(driver)
    driver.find_element(By.ID, "keteranganForm").click()
    driver.find_element(By.ID, "keteranganForm").send_keys(Keterangan)
    Log.info('Input Keterangan')

@mark.fixture_test()
def test_12_ClickButtonSubmit():
    sleep(driver)
    driver.find_element(By.CSS_SELECTOR, "#submitButton path").click()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, 'searchButton')))
    Log.info('Click Button Submit')

@mark.fixture_test()
def test_exit():
    quit(driver)

