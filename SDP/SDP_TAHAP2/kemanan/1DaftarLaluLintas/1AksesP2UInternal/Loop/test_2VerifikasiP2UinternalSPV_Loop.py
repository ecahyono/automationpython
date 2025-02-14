from distutils.archive_util import make_archive
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium.webdriver.firefox.options import Options

options = Options()
options.add_argument('--disable-blink-features=AutomationControlled')
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUAT"))

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))


from Settings.setup import initDriver, loadDataPath, quit, sleep
from Settings.login import loginOperatorSumedang, Op_Keamanan_p2u, SpvRutanBdg
from Settings.Page.keamanan import p2uinternal

import logging
Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('Log2_P2UinternalSpv.log', mode="w")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)


sheetrangeIndex = wb['P2U_Internal']

@mark.fixture_test()
def test_1_setupOS():
    global driver, pathData
    driver = initDriver()
    pathData = loadDataPath()
    Log.info('Setup Os')

@mark.fixture_test()
def test_2_login():
    SpvRutanBdg(driver)
    Log.info('Login Super Visor P2U')


@mark.fixture_test()
def test_3_Input():
    sleep(driver)
    p2uinternal(driver)
    print('.')
    Log.info('Akses halaman Daftar Lalu Lintas P2U Internal')
    attach(data=driver.get_screenshot_as_png())


    i = 2
        
    while i <= len(sheetrangeIndex['A']):
    
        NamaInput                                 = sheetrangeIndex['A'+str(i)].value
        Keterangan                                = sheetrangeIndex['B'+str(i)].value
            
        try:
            sleep(driver)
            
            print('Pilih Dropdown Nama')
            sleep(driver)
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))

            driver.find_element(By.ID, 'statusColumn').click()
            driver.find_element(By.XPATH, "//li[contains(.,\'Dalam Proses\')]").click()
            Log.info('Klik Status Dalam Proses')

            sleep(driver)
            driver.find_element(By.XPATH, '//*[@id="kataKunci"]').clear()
            driver.find_element(By.XPATH, '//*[@id="kataKunci"]').send_keys(NamaInput)
            Log.info('Search Nama WBP')

            driver.find_element(By.ID, 'searchButton' ).click()
            Log.info('Click Button Search')
            sleep(driver)

            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchButton"]')))
            driver.find_element(By.CSS_SELECTOR, ".w-6 > path").click()
            Log.info('Click Button Verifikasi')
            
            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'keterangan')))
            driver.find_element(By.ID, 'keterangan').send_keys(Keterangan)
            Log.info('Input Keterangan')

            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, 'Izinkan')))
            driver.find_element(By.ID, 'Izinkan').click()
            Log.info('click button izinkan')
            driver.execute_script("window.scrollTo(0,527)") 

            WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(.,'Ya')]")))
            driver.find_element(By.XPATH, "//span[contains(.,'Ya')]").click()
            Log.info ('Click button ya ')
            sleep(driver)


        except TimeoutException:
            print("ERRROR")
            pass
                
        sleep(driver)
        i = i + 1
    print('DONE')

def teardown():
    quit(driver)
        

  
        






