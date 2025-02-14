from openpyxl import Workbook
from faker import Faker
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
import platform
from pytest import mark
import time
from pytest_html_reporter import attach
import pyautogui
from datetime import datetime
import pytest
from selenium import webdriver

import sys
from os import environ, path
from dotenv import load_dotenv
load_dotenv()
from openpyxl import load_workbook


if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR"))
    wb = load_workbook(environ.get("data"))
    file_path = environ.get("fakerGiatja")

elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))
    wb = load_workbook(environ.get("KeamananUATWin"))
import random

from Settings.SetupAll import *
from Settings.loginPembinaan import *
from Settings.Page.Pembinaan import *
import random
import logging

sheetrangeIndex = wb['listWbpSumedang']

random1 = random.randint(1,170)
random2= random.randint(1,170)
random3 = random.randint(1,170)

NamaInput1                                 = sheetrangeIndex['A'+str(random1)].value
NamaInput2                                 = sheetrangeIndex['B'+str(random2)].value
NamaInput3                                 = sheetrangeIndex['C'+str(random3)].value
time.sleep(3)

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'GiatjaTambah'

fake = Faker('id_ID')

KegiatanKerja   = ['Manufaktur','Jasa']
KegiatanKerja1   = ['bidang1-opt0','bidang1-opt1','bidang1-opt2']
KegiatanKerja2   = ['bidang1-opt0','bidang1-opt1','bidang1-opt2']

Jeniskegiatan   = ['jenisKegiatan0','jenisKegiatan1','jenisKegiatan2']


Skalakegiatan           = ['skala0','skala1','skala2']
area                    = ['area0','area1','area2']
sarana                  = ['sarana0']
prasarana               = ['prasaranaundefined-0']
mitra                   = ['mitra0']
checkbox                = ['.el-table__row:nth-child(1) .el-checkbox__inner','.el-table__row:nth-child(2) .el-checkbox__inner','.el-table__row:nth-child(3) .el-checkbox__inner']
checkbox1               = ['.el-table__row:nth-child(4) .el-checkbox__inner','.el-table__row:nth-child(5) .el-checkbox__inner','.el-table__row:nth-child(6) .el-checkbox__inner']
checkbox2               = ['.el-table__row:nth-child(7) .el-checkbox__inner','.el-table__row:nth-child(8) .el-checkbox__inner','.el-table__row:nth-child(9) .el-checkbox__inner']
idjenis                 = ['jenis0']
satuan                  = ['satuan0','satuan1','satuan3','satuan4','satuan5']
fakeMonth               = ['09']
fakeMonthRevisi         = ['04']
akunSetor                                               = ['pendapatan','nihil']



for i in range(1):
    KegiatanKerjaFaker                = random.choice(KegiatanKerja)
    KegiatanKerjaFaker1               = random.choice(KegiatanKerja1)
    KegiatanKerjaFaker2               = random.choice(KegiatanKerja2)
    
    JeniskegiatanFaker                = random.choice(Jeniskegiatan)
    namaKegiatanFaker                 = fake.text(max_nb_chars=7)
    SkalakegiatanFaker                = random.choice(Skalakegiatan)
    tanggalAwalKegiatanFaker          = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    tanggalAkhirKegiatanFaker         = fake.date_between(start_date='today', end_date='today').strftime('%d/%m/%Y')
    areaFaker                         = random.choice(area)
    lokasiKegiatanFaker               = fake.address()
    luasLokasiKegiatanFaker           = fake.random_int(min=10, max=1000)
    jumlahRuangKegiatanFaker          = fake.random_int(min=1, max=10)
    saranaFaker                       = random.choice(sarana)
    prasaranaFaker                    = random.choice(prasarana)
    mitraFaker                        = random.choice(mitra)
    ketereranganFaker                 = fake.text(max_nb_chars=100)
    jumlahPesertaFaker                = fake.random_int(min=1, max=3)
    PesertaRandomFaker1               = random.choice(checkbox)
    PesertaRandomFaker2               = random.choice(checkbox1)
    PesertaRandomFaker3               = random.choice(checkbox2)
    idjenisFaker                      = random.choice(idjenis)
    namaProdukFaker                   = fake.text(max_nb_chars=7)
    jumlahprodukFaker                 = fake.random_int(min=1, max=3)
    satuanFaker                       = random.choice(satuan)
    HargaFaker                        = fake.random_int(min=10000, max=1000000)
    lamaPengerjaanFaker               = fake.random_int(min=1, max=7)

    TanggalSetorFaker                                   = random.choice(fakeMonth)
    TanggalSetorFakerRevisi                                   = random.choice(fakeMonthRevisi)

    NilaiFaker                                          = fake.random_int(min=100000, max=1000000)

    # TanggalSetorFaker                                   = fake.date_between(start_date='today', end_date='today').strftime('%m')
    akunSetorFaker                                      = random.choice(akunSetor)
    keteranganFaker                                     = fake.text(max_nb_chars=255)
    
    




    worksheet.append([
        KegiatanKerjaFaker,
        JeniskegiatanFaker,
        namaKegiatanFaker,
        SkalakegiatanFaker,
        tanggalAwalKegiatanFaker,
        tanggalAkhirKegiatanFaker,
        areaFaker,
        lokasiKegiatanFaker,
        luasLokasiKegiatanFaker,
        jumlahRuangKegiatanFaker,
        saranaFaker,
        prasaranaFaker,
        mitraFaker,
        ketereranganFaker,
        jumlahPesertaFaker,
        PesertaRandomFaker1,
        PesertaRandomFaker2,
        PesertaRandomFaker3,
        idjenisFaker,
        namaProdukFaker,
        jumlahprodukFaker,
        satuanFaker,
        HargaFaker,
        lamaPengerjaanFaker
        ])
workbook.save(file_path)

workbook = load_workbook(filename=file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=1, values_only=True):
    KegiatanKerjaFkr                            = row[0]
    JeniskegiatanFkr                            = row[1]
    namaKegiatanFkr                             = row[2]
    SkalakegiatanFkr                            = row[3]
    tanggalAwalKegiatanFkr                      = row[4]
    tanggalAkhirKegiatanFkr                     = row[5]
    areaFkr                                     = row[6]
    lokasiKegiatanFkr                           = row[7]
    luasLokasiKegiatanFkr                       = row[8]
    jumlahRuangKegiatanFkr                      = row[9]
    saranaFkr                                   = row[10]
    prasaranaFkr                                = row[11]
    mitraFkr                                    = row[12]
    ketereranganFkr                             = row[13]
    jumlahPesertaFkr                            = row[14]
    PesertaRandomFkr                            = row[15]
    PesertaRandomFkr2                           = row[16]
    PesertaRandomFkr3                           = row[17]
    idjenisFkr                                  = row[18]
    namaProdukFkr                               = row[19]
    jumlahprodukFkr                             = row[20]
    satuanFkr                                   = row[21]
    HargaFkr                                    = row[22]
    lamaPengerjaanFkr                           = row[23]


global driver, pathData
driver = initDriver()
pathData = loadDataPath()      

