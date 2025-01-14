from turtle import rt
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import os
import pyautogui
import pytest
import time
import platform
import subprocess

from pathlib import Path

@pytest.fixture()
def test_setup():
    global driver
    global wb
    swin = Service(r'C:/Users/user/Documents/TRCH/chromedriver.exe')
    smac = Service('/Users/will/Downloads/chromedriver')

    if platform.system() == 'Darwin':
        driver = webdriver.Chrome(service=smac)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max
        wb = load_workbook(filename=r"/Users/will/Documents/work/Automationpython/Filexel/Registrasi.xlsx")   
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
    elif platform.system() == 'Windows':
        driver = webdriver.Chrome(service=swin)
        # url = "http://kumbang.torche.id:32400/"
        url = "http://192.168.2.11:32400/"
        
        driver.get(url)
        # seting windows nya jadi max   
        wb = load_workbook(filename=r"C:\Users\user\Documents\TRCH\Automationpython\Filexel\Registrasi.xlsx")
        driver.maximize_window()
        driver.implicitly_wait(5)
        yield
        driver.close()
        driver.quit()
        

def test_negara(test_setup):
    sheetrange = wb['NamaNegara']
    driver.find_element(By.XPATH, "//div/span").click()
    # ini masuk ke form input username
    driver.find_element(By.ID, "username").click()
    # masukin input username
    driver.find_element(By.ID, "username").send_keys("wildan")
    # masukin input password
    driver.find_element(By.ID, "password").send_keys("wildan")
    # click button login
    driver.find_element(By.ID, "kc-login").click()
    time.sleep(2)
    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[8]/div")
    time.sleep(2)
    actions = ActionChains(driver)
    actions.move_to_element(element).perform()
    time.sleep(2)

    element = driver.find_element(By.XPATH, "//*[@id=\"app\"]/div/nav/ul/li[8]/div")

    actions = ActionChains(driver)
    actions.move_to_element(element).perform()


    element2 = driver.find_element(By.XPATH, "//div[10]/div/ul/li/div")

    actions2 = ActionChains(driver)
    actions2.move_to_element(element2).perform()
    driver.find_element(By.LINK_TEXT, "Negara").click()

    i = 2

    # nge baca mulai dari tabel A
    while i <= len(sheetrange['A']):
        # deklarasi bahwa NIP itu ada di A 
        NamaNegara = sheetrange['A'+str(i)].value
        driver.find_element(By.CSS_SELECTOR, ".is-plain:nth-child(2)").click()
        try:      
            time.sleep(2)
            driver.find_element(By.XPATH, "//input[@type=\'text\']").click()
            driver.find_element(By.XPATH, "//input[@type=\'text\']").send_keys(NamaNegara)
            driver.find_element(By.CSS_SELECTOR, ".el-form-item__content > .w-full").click()
            driver.find_element(By.XPATH, "//button[contains(.,\'Simpan\')]").click()
    
        except TimeoutException:
            print("MASIH ADA ERROR, CEK LAGI PAK WIL")
            pass
        time.sleep(5)
        i = i + 1
    print("DONE PAK WILDAN, SEBATS DULU")
